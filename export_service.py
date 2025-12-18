import re
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import traceback
from core import get_model_config


# ============================================================
# PATHS FROM CONFIG
# ============================================================

def _get_paths():
    config = get_model_config()

    excel_path = Path(config.OUTPUT_EXCEL_PATH)
    json_backup_path = excel_path.with_suffix(".json")

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    return excel_path, json_backup_path


EXCEL_COLUMNS = [
    "index",
    "days_count",
    "region",
    "deity",
    "yatra_specifics",
    "pace",
    "transport_preference",
    "additional_instructions",
    "specific_interests",
    "generated_itinerary",
]


# ============================================================
# UTIL: MARKDOWN CLEANER
# ============================================================

def __markdown_to_text(data: str) -> str:
    stuff_to_remove = [r'### ', r'####', r'##', r'#', r' - ', r'- ', r'---', r'**', r'***', r'  ']
    for s in stuff_to_remove:
        data = re.sub(re.escape(s), '', data)
    return data


# ============================================================
# JSON BACKUP
# ============================================================

def write_json_backup(filters: dict, output: str):
    excel_path, backup_path = _get_paths()

    entry = {**filters, "generated_itinerary": output}

    if backup_path.exists():
        try:
            data = json.loads(backup_path.read_text())
        except json.JSONDecodeError:
            data = []
    else:
        data = []

    data.append(entry)
    backup_path.write_text(json.dumps(data, indent=4))


# ============================================================
# EXCEL INITIALIZATION
# ============================================================

def _initialize_excel():
    excel_path, _ = _get_paths()

    if not excel_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(EXCEL_COLUMNS)
        _format_header(ws)
        wb.save(excel_path)
        return wb, ws

    wb = load_workbook(excel_path)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.create_sheet("Data")

    # Ensure header exists
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(EXCEL_COLUMNS)
        _format_header(ws)

    return wb, ws


# ============================================================
# FORMATTERS
# ============================================================

def _format_header(ws):
    """Make headers bold + centered."""
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _format_row(ws, row_number: int):
    """Center-align every column for readability."""
    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        ws.cell(row=row_number, column=col_idx).alignment = Alignment(
            horizontal="center", vertical="center"
        )


def _auto_adjust_width(ws):
    """Adjust column widths based on longest cell value."""
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)

        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_len + 3


# ============================================================
# WRITE TO EXCEL + BACKUP
# ============================================================

def write_to_excel_with_sync(filters: dict, output: str) -> bool:
    """
    Try writing the entry to Excel.
    Always writes JSON backup first.
    Returns True if written to Excel, False if only saved to JSON (e.g. Excel open).
    """
    output = __markdown_to_text(output)
    excel_path, backup_path = _get_paths()

    # 1) always write JSON backup first
    write_json_backup(filters, output)

    try:
        wb, ws = _initialize_excel()

        # Determine next index (use max_row to preserve continuity)
        next_index = ws.max_row if ws.max_row >= 2 else 1

        row_number = ws.max_row + 1
        ws.cell(row=row_number, column=1, value=next_index)
        ws.cell(row=row_number, column=2, value=filters.get("days_count"))
        ws.cell(row=row_number, column=3, value=filters.get("region"))
        ws.cell(row=row_number, column=4, value=filters.get("deity"))
        ws.cell(row=row_number, column=5, value=filters.get("yatra_specifics"))
        ws.cell(row=row_number, column=6, value=filters.get("pace"))
        ws.cell(row=row_number, column=7, value=filters.get("transport_preference"))
        ws.cell(row=row_number, column=8, value=filters.get("additional_instructions"))
        ws.cell(row=row_number, column=9, value=filters.get("specific_interests"))

        ws.cell(row=row_number, column=10, value=str(output))

        _format_row(ws, row_number)
        _auto_adjust_width(ws)

        wb.save(excel_path)

        print(f"[export_service] Excel write successful. Path: {excel_path}")
        return True

    except PermissionError:
        print(f"[export_service] PermissionError: Excel file appears to be open. Path: {excel_path}")
        return False

    except Exception as exc:
        print(f"[export_service] Unexpected error while writing to Excel: {exc}")
        traceback.print_exc()
        return False


def sync_missing_entries() -> int:
    """
    Sync missing JSON entries into Excel.
    Removes entries from JSON that were successfully written.
    Returns number of rows added.
    """
    excel_path, backup_path = _get_paths()

    if not backup_path.exists():
        print("[export_service] No backup file found.")
        return 0

    try:
        backup_data = json.loads(backup_path.read_text())
    except json.JSONDecodeError:
        print("[export_service] JSON corrupted — resetting.")
        backup_path.write_text("[]")
        return 0

    if not backup_data:
        print("[export_service] Backup is empty.")
        return 0

    try:
        wb, ws = _initialize_excel()
    except PermissionError:
        print(f"[export_service] Cannot open Excel file (PermissionError). Path: {excel_path}")
        return 0

    # Build a set of existing rows (excluding index)
    existing_tuples = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_tuples.add(tuple(row[1:]))

    remaining_entries = []
    added = 0

    for entry in backup_data:
        row_tuple = (
            entry.get("days_count"),
            entry.get("region"),
            entry.get("deity"),
            entry.get("yatra_specifics"),
            entry.get("pace"),
            entry.get("transport_preference"),
            entry.get("additional_instructions"),
            entry.get("specific_interests"),
            entry.get("generated_itinerary"),
        )

        if row_tuple in existing_tuples:
            # already written -> skip and DO NOT keep in JSON
            continue

        try:
            row_number = ws.max_row + 1
            ws.cell(row=row_number, column=1, value=ws.max_row)  # index
            for col_idx, value in enumerate(row_tuple, start=2):
                if col_idx == 10:
                    ws.cell(row=row_number, column=col_idx, value=str(value))
                else:
                    ws.cell(row=row_number, column=col_idx, value=value)

            _format_row(ws, row_number)
            existing_tuples.add(row_tuple)
            added += 1

        except PermissionError:
            print("[export_service] Excel locked during sync — aborting.")
            remaining_entries.append(entry)
            break
        except Exception as e:
            print(f"[export_service] Error while syncing entry: {e}")
            traceback.print_exc()
            remaining_entries.append(entry)
            continue

    if added > 0:
        _auto_adjust_width(ws)
        wb.save(excel_path)

    backup_path.write_text(json.dumps(remaining_entries, indent=4))

    print(f"[export_service] Sync complete. Added {added} rows. Remaining in JSON: {len(remaining_entries)}")
    return added

# ============================================================
# PUBLIC METHOD
# ============================================================

def export_itinerary(filters: dict, itinerary_text: str):
    write_to_excel_with_sync(filters, itinerary_text)


__all__ = [
    "export_itinerary",
    "sync_missing_entries",
    "write_json_backup",
    "write_to_excel_with_sync",
]
